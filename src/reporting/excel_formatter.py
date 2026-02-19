"""
Excel formatter for Swaption Vol Table - Nomura-style formatting
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from typing import Optional
from pathlib import Path


def format_swaption_vol_table_excel(
    table: pd.DataFrame,
    as_of_date: date,
    output_file: Optional[str] = None
) -> str:
    """
    Format Swaption Vol Table as Excel file with Nomura-style formatting
    
    Args:
        table: Swaption vol table DataFrame
        as_of_date: Date of the table
        output_file: Output filename (default: swaption_vol_table_YYYY-MM-DD.xlsx)
        
    Returns:
        Path to created Excel file
    """
    if output_file is None:
        # Save to outputs/tables/ directory
        output_dir = Path(__file__).parent.parent.parent / "outputs" / "tables"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = str(output_dir / f'swaption_vol_table_{as_of_date}.xlsx')
    else:
        # If path provided, ensure it's a string
        output_file = str(output_file)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Swaption Vol Table"
    
    # Define styles
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    section_header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    mover_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Dark gray
    mover_font = Font(bold=True, color="FFFFFF")  # White text for dark background
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    number_format = "#,##0.00"
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Start row
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:V{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = f"Vol Monitor – Swaption Vol Table"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    row += 1
    
    # Subtitle with date
    ws.merge_cells(f'A{row}:V{row}')
    subtitle_cell = ws[f'A{row}']
    subtitle_cell.value = f"As of {as_of_date}"
    subtitle_cell.font = Font(size=11)
    subtitle_cell.alignment = center_align
    row += 2
    
    # Main header row
    ws.merge_cells(f'A{row}:A{row+1}')  # Term/Tenor spans 2 rows
    ws[f'A{row}'].value = "Term/Tenor"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    
    # Section 1: Implied Vol (Annualized)
    col = 2
    ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+5)}{row}')
    ws[f'{get_column_letter(col)}{row}'].value = "Implied Basis Point Volatility (Annualized)"
    ws[f'{get_column_letter(col)}{row}'].font = header_font
    ws[f'{get_column_letter(col)}{row}'].fill = section_header_fill
    ws[f'{get_column_letter(col)}{row}'].alignment = center_align
    ws[f'{get_column_letter(col)}{row}'].border = thin_border
    
    ann_cols = ['Current', '1d Chg', '1w Chg', '1m Chg', '20d High', '20d Low']
    for i, col_name in enumerate(ann_cols):
        cell = ws[f'{get_column_letter(col+i)}{row+1}']
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Section 2: Implied Vol (Daily)
    col = 8
    ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+5)}{row}')
    ws[f'{get_column_letter(col)}{row}'].value = "Implied Basis Point Volatility (Daily)"
    ws[f'{get_column_letter(col)}{row}'].font = header_font
    ws[f'{get_column_letter(col)}{row}'].fill = section_header_fill
    ws[f'{get_column_letter(col)}{row}'].alignment = center_align
    ws[f'{get_column_letter(col)}{row}'].border = thin_border
    
    for i, col_name in enumerate(ann_cols):
        cell = ws[f'{get_column_letter(col+i)}{row+1}']
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Section 3: Realized Vol (Daily)
    col = 14
    ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+5)}{row}')
    ws[f'{get_column_letter(col)}{row}'].value = "Realized Basis Point Volatility (Daily)"
    ws[f'{get_column_letter(col)}{row}'].font = header_font
    ws[f'{get_column_letter(col)}{row}'].fill = section_header_fill
    ws[f'{get_column_letter(col)}{row}'].alignment = center_align
    ws[f'{get_column_letter(col)}{row}'].border = thin_border
    
    realized_cols = ['10d', '20d', '60d', '90d', '120d', '180d']
    for i, col_name in enumerate(realized_cols):
        cell = ws[f'{get_column_letter(col+i)}{row+1}']
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 2
    
    # Data rows
    for idx, table_row in table.iterrows():
        # Term/Tenor
        cell = ws[f'A{row}']
        cell.value = table_row['term_tenor']
        cell.alignment = center_align
        cell.border = thin_border
        
        # Implied Vol (Annualized)
        ann_data = [
            table_row['implied_vol_ann_current'],
            table_row['implied_vol_ann_1d_chg'],
            table_row['implied_vol_ann_1w_chg'],
            table_row['implied_vol_ann_1m_chg'],
            table_row['implied_vol_ann_20d_high'],
            table_row['implied_vol_ann_20d_low'],
        ]
        
        for i, val in enumerate(ann_data):
            cell = ws[f'{get_column_letter(2+i)}{row}']
            if pd.isna(val):
                cell.value = ""
            else:
                # Format negative values with parentheses
                if val < 0:
                    cell.value = f"({abs(val):.2f})"
                else:
                    cell.value = val
                    cell.number_format = number_format
            cell.alignment = right_align
            cell.border = thin_border
            
            # Highlight if largest mover
            col_name = ann_cols[i]
            if col_name == '1d Chg' and table_row.get('is_largest_1d_mover', False):
                cell.fill = mover_fill
                cell.font = mover_font
            elif col_name == '1w Chg' and table_row.get('is_largest_1w_mover', False):
                cell.fill = mover_fill
                cell.font = mover_font
            elif col_name == '1m Chg' and table_row.get('is_largest_1m_mover', False):
                cell.fill = mover_fill
                cell.font = mover_font
        
        # Implied Vol (Daily)
        daily_data = [
            table_row['implied_vol_daily_current'],
            table_row['implied_vol_daily_1d_chg'],
            table_row['implied_vol_daily_1w_chg'],
            table_row['implied_vol_daily_1m_chg'],
            table_row['implied_vol_daily_20d_high'],
            table_row['implied_vol_daily_20d_low'],
        ]
        
        for i, val in enumerate(daily_data):
            cell = ws[f'{get_column_letter(8+i)}{row}']
            if pd.isna(val):
                cell.value = ""
            else:
                # Format negative values with parentheses
                if val < 0:
                    cell.value = f"({abs(val):.2f})"
                else:
                    cell.value = val
                    cell.number_format = number_format
            cell.alignment = right_align
            cell.border = thin_border
            
            # Highlight if largest mover
            col_name = ann_cols[i]
            if col_name == '1d Chg' and table_row.get('is_largest_1d_mover', False):
                cell.fill = mover_fill
                cell.font = mover_font
            elif col_name == '1w Chg' and table_row.get('is_largest_1w_mover', False):
                cell.fill = mover_fill
                cell.font = mover_font
            elif col_name == '1m Chg' and table_row.get('is_largest_1m_mover', False):
                cell.fill = mover_fill
                cell.font = mover_font
        
        # Realized Vol (Daily)
        realized_data = [
            table_row.get('realized_vol_10d', np.nan),
            table_row.get('realized_vol_20d', np.nan),
            table_row.get('realized_vol_60d', np.nan),
            table_row.get('realized_vol_90d', np.nan),
            table_row.get('realized_vol_120d', np.nan),
            table_row.get('realized_vol_180d', np.nan),
        ]
        
        for i, val in enumerate(realized_data):
            cell = ws[f'{get_column_letter(14+i)}{row}']
            if pd.isna(val):
                cell.value = ""
            else:
                cell.value = val
                cell.number_format = number_format
            cell.alignment = right_align
            cell.border = thin_border
        
        row += 1
    
    # Add notes
    row += 2
    ws.merge_cells(f'A{row}:V{row}')
    notes_cell = ws[f'A{row}']
    notes_cell.value = "Notes on Color Coding"
    notes_cell.font = Font(bold=True, size=10)
    notes_cell.alignment = Alignment(horizontal="left")
    row += 1
    
    ws.merge_cells(f'A{row}:V{row}')
    notes_text = ws[f'A{row}']
    notes_text.value = "Largest movers (1-day largest movers over 2 weeks; 1-week largest movers over 1 month; 1-month largest movers over 6 months)"
    notes_text.font = Font(size=9)
    notes_text.alignment = Alignment(horizontal="left")
    
    # Add source
    row += 1
    ws.merge_cells(f'A{row}:V{row}')
    source_cell = ws[f'A{row}']
    source_cell.value = "Source: VolCube420, SOFR Swap Rates"
    source_cell.font = Font(size=9, italic=True)
    source_cell.alignment = Alignment(horizontal="left")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Term/Tenor
    for col in range(2, 21):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Save
    wb.save(output_file)
    return output_file


def format_negative_values(value: float) -> str:
    """
    Format negative values with parentheses (accounting style)
    
    Args:
        value: Numeric value
        
    Returns:
        Formatted string
    """
    if pd.isna(value):
        return ""
    if value < 0:
        return f"({abs(value):.2f})"
    return f"{value:.2f}"
